#!/usr/bin/env python3
"""
MLS Excel Formatter - Creating Insanely Great Spreadsheets

This module takes extracted MLS data and creates a professional Excel file
that looks like it was designed by hand.

Philosophy: Perfect is the only acceptable standard.

Author: Claude Code
Version: 1.0.0
Date: 2025-11-06
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
from datetime import datetime
from zoneinfo import ZoneInfo


# Perfect Column Order - by decision importance
COLUMN_ORDER = [
    # Critical decision data first
    'is_subject',
    'address',
    'unit',
    'available_sf',
    'net_asking_rent',
    'tmi',
    'gross_rent',

    # Important property characteristics
    'clear_height_ft',
    'building_age_years',
    'class',
    'parking_ratio',
    'pct_office_space',

    # Secondary characteristics
    'shipping_doors_tl',
    'shipping_doors_di',
    'power_amps',
    'bay_depth_ft',
    'lot_size_acres',

    # Tertiary characteristics
    'hvac_coverage',
    'sprinkler_type',
    'rail_access',
    'crane',
    'occupancy_status',
    'trailer_parking',
    'secure_shipping',
    'excess_land',
    'grade_level_doors',
    'zoning',

    # Metadata
    'availability_date',
    'days_on_market',
    'mls_number',
    'broker_name',
    'client_remarks',
    'reported_market',
    'report_generated_at',
    'source_pdf',
    'year_built'  # Keep for reference even though we show building_age
]


# Perfect Column Headers - human-readable
COLUMN_HEADERS = {
    'is_subject': 'Subject Property',
    'address': 'Address',
    'unit': 'Unit',
    'available_sf': 'Available SF',
    'net_asking_rent': 'Net Rent ($/SF)',
    'tmi': 'TMI ($/SF)',
    'gross_rent': 'Gross Rent ($/SF)',
    'clear_height_ft': 'Clear Height (ft)',
    'building_age_years': 'Building Age (yrs)',
    'class': 'Class',
    'parking_ratio': 'Parking Ratio',
    'pct_office_space': '% Office',
    'shipping_doors_tl': 'Ship Doors (TL)',
    'shipping_doors_di': 'Ship Doors (DI)',
    'power_amps': 'Power (amps)',
    'bay_depth_ft': 'Bay Depth (ft)',
    'lot_size_acres': 'Lot Size (acres)',
    'hvac_coverage': 'HVAC',
    'sprinkler_type': 'Sprinkler',
    'rail_access': 'Rail Access',
    'crane': 'Crane',
    'occupancy_status': 'Occupancy',
    'trailer_parking': 'Trailer Parking',
    'secure_shipping': 'Secure Shipping',
    'excess_land': 'Excess Land',
    'grade_level_doors': 'Grade Level Doors',
    'zoning': 'Zoning',
    'availability_date': 'Availability Date',
    'days_on_market': 'Days on Market',
    'mls_number': 'MLS#',
    'broker_name': 'Broker',
    'client_remarks': 'Remarks',
    'reported_market': 'Market',
    'report_generated_at': 'Report Date',
    'source_pdf': 'Source PDF',
    'year_built': 'Year Built'
}


def create_perfect_excel(properties: List[Dict[str, Any]], output_path: str) -> str:
    """
    Create an insanely great Excel file from MLS property data.

    This function embodies the philosophy: perfect is the only acceptable standard.
    Every detail matters - colors, fonts, spacing, alignment, number formatting.

    Args:
        properties: List of property dictionaries with standardized fields
        output_path: Full path to output .xlsx file

    Returns:
        Path to created Excel file

    Raises:
        ValueError: If properties list is empty
        KeyError: If required fields are missing
    """
    if not properties:
        raise ValueError("Cannot create Excel file from empty properties list")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "MLS Properties"

    # Write headers with perfect styling
    _write_headers(ws)

    # Write data rows with perfect formatting
    _write_data_rows(ws, properties)

    # Apply perfect column widths
    _apply_perfect_widths(ws)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Enable auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Save with perfect settings
    wb.save(output_path)

    print(f"✅ Created perfect Excel file: {output_path}")
    return output_path


def _write_headers(ws):
    """
    Write header row with perfect styling.

    Dark blue background, white bold text, centered.
    This is the first thing users see - it must be perfect.
    """
    # Perfect header styling
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_border = Border(
        bottom=Side(style='medium', color='FFFFFF')
    )

    for col_idx, field in enumerate(COLUMN_ORDER, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = COLUMN_HEADERS.get(field, field.replace('_', ' ').title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    # Perfect row height for headers
    ws.row_dimensions[1].height = 30


def _write_data_rows(ws, properties: List[Dict[str, Any]]):
    """
    Write data rows with perfect formatting and highlighting.

    Subject property gets bright yellow background - impossible to miss.
    Alternating row colors for easy reading.
    Perfect number formatting for currency, percentages, integers.
    """
    # Perfect styling for data rows
    normal_font = Font(name='Calibri', size=10)
    subject_font = Font(name='Calibri', size=10, bold=True)

    # Alternating row fills
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    gray_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    subject_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Bright yellow

    # Perfect alignment
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    center_align = Alignment(horizontal='center', vertical='center')

    # Subtle borders
    thin_border = Border(
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0'),
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0')
    )

    for row_idx, prop in enumerate(properties, start=2):
        is_subject = prop.get('is_subject', False)

        # Choose fill based on subject/alternating pattern
        if is_subject:
            row_fill = subject_fill
            row_font = subject_font
        else:
            row_fill = white_fill if row_idx % 2 == 0 else gray_fill
            row_font = normal_font

        for col_idx, field in enumerate(COLUMN_ORDER, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = prop.get(field)

            # Write value with perfect type handling
            cell.value = _format_cell_value(field, value)
            cell.font = row_font
            cell.fill = row_fill
            cell.border = thin_border

            # Perfect alignment based on data type
            if field in ['is_subject', 'class', 'hvac_coverage', 'sprinkler_type', 'occupancy_status']:
                cell.alignment = center_align
            elif isinstance(value, (int, float)):
                cell.alignment = right_align
            else:
                cell.alignment = left_align

            # Perfect number formatting
            _apply_number_format(cell, field)


def _format_cell_value(field: str, value: Any) -> Any:
    """
    Format cell value for perfect display.

    Converts Python types to Excel-friendly formats.
    """
    if value is None:
        return ""

    # Boolean fields
    if field in ['is_subject', 'rail_access', 'crane', 'trailer_parking', 'secure_shipping', 'excess_land']:
        return "YES" if value else "NO"

    # Class field (convert 1/2/3 to A/B/C)
    if field == 'class':
        class_map = {1: 'A', 2: 'B', 3: 'C'}
        return class_map.get(value, value)

    # HVAC field (ordinal to text)
    if field == 'hvac_coverage':
        hvac_map = {1: 'Y', 2: 'Partial', 3: 'N'}
        return hvac_map.get(value, value)

    # Sprinkler field (ordinal to text)
    if field == 'sprinkler_type':
        sprinkler_map = {1: 'ESFR', 2: 'Standard', 3: 'None'}
        return sprinkler_map.get(value, value)

    # Occupancy field (ordinal to text)
    if field == 'occupancy_status':
        occupancy_map = {1: 'Vacant', 2: 'Occupied'}
        return occupancy_map.get(value, value)

    return value


def _apply_number_format(cell, field: str):
    """
    Apply perfect number formatting based on field type.

    Currency, percentages, integers - all formatted professionally.
    """
    # Currency fields ($/SF)
    if field in ['net_asking_rent', 'tmi', 'gross_rent']:
        cell.number_format = '$#,##0.00'

    # Percentage fields
    elif field == 'pct_office_space':
        cell.number_format = '0.0%'

    # Integer fields with thousands separator
    elif field in ['available_sf', 'power_amps', 'days_on_market']:
        cell.number_format = '#,##0'

    # Decimal fields (1 decimal place)
    elif field in ['clear_height_ft', 'parking_ratio', 'building_age_years', 'bay_depth_ft', 'lot_size_acres']:
        cell.number_format = '0.0'

    # Integer fields (no separator)
    elif field in ['shipping_doors_tl', 'shipping_doors_di', 'grade_level_doors', 'year_built']:
        cell.number_format = '0'


def _apply_perfect_widths(ws):
    """
    Apply perfect column widths.

    Auto-size based on content, with sensible min/max constraints.
    No endless wide columns. No cramped narrow columns.
    """
    for col_idx, field in enumerate(COLUMN_ORDER, start=1):
        column_letter = get_column_letter(col_idx)

        # Calculate ideal width based on content
        max_length = 0
        for row in ws[column_letter]:
            try:
                cell_value = str(row.value) if row.value else ""
                max_length = max(max_length, len(cell_value))
            except:
                pass

        # Perfect width constraints
        min_width = 10
        max_width = 50
        ideal_width = min(max(max_length + 2, min_width), max_width)

        ws.column_dimensions[column_letter].width = ideal_width


def generate_filename(market: str = "properties") -> str:
    """
    Generate perfect filename with Eastern Time timestamp.

    Format: YYYY-MM-DD_HHMMSS_mls_extraction_<market>.xlsx

    Args:
        market: Market name (auto-detected from PDF or default)

    Returns:
        Filename string
    """
    # Eastern Time
    et = ZoneInfo('America/New_York')
    timestamp = datetime.now(et).strftime('%Y-%m-%d_%H%M%S')

    # Clean market name (remove spaces, special chars)
    clean_market = market.lower().replace(' ', '_').replace('-', '_')
    clean_market = ''.join(c for c in clean_market if c.isalnum() or c == '_')

    return f"{timestamp}_mls_extraction_{clean_market}.xlsx"


# CLI interface
if __name__ == '__main__':
    import sys
    import json
    import argparse

    parser = argparse.ArgumentParser(
        description='Create insanely great Excel files from MLS extraction data'
    )
    parser.add_argument('input_json', help='Path to JSON file with extracted properties')
    parser.add_argument('output_excel', nargs='?', help='Path to output Excel file (optional, will auto-generate if not provided)')

    args = parser.parse_args()

    # Load JSON data
    try:
        with open(args.input_json, 'r') as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"❌ Error: Input file not found: {args.input_json}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"❌ Error: Invalid JSON: {e}")
        sys.exit(1)

    # Extract properties list
    properties = data.get('properties', [])
    if not properties:
        print("❌ Error: No properties found in JSON data")
        sys.exit(1)

    # Generate output path if not provided
    if args.output_excel:
        output_path = args.output_excel
    else:
        # Auto-generate based on market name
        import os
        market = data.get('market', 'properties')
        filename = generate_filename(market)
        output_dir = os.path.join(os.path.dirname(args.input_json), '')
        output_path = os.path.join(output_dir, filename)

    # Create perfect Excel
    try:
        create_perfect_excel(properties, output_path)
        print(f"\n✅ SUCCESS!")
        print(f"📊 Extracted {len(properties)} properties")
        print(f"📁 Excel file: {output_path}")

        # Report subject property
        subject = next((p for p in properties if p.get('is_subject')), None)
        if subject:
            print(f"🎯 Subject property: {subject.get('address', 'Unknown')}")

    except Exception as e:
        print(f"❌ Error creating Excel: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
